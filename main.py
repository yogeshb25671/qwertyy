# execrsices in this projet
# 1)list each company with their respective product count
# 3)list the products with inventory less than 10
# 2)list each company with respective total inventory Value
# 4)write to spreadsheet:calculate and write inventory value for each product into spreadsheet


import openpyxl
inv_file=openpyxl.load_workbook("inventory.xlsx") #this refers to entire excel file 

product_list=inv_file["Sheet1"]  #this refer to sheet1


product_per_supplier={} #1)
total_value_per_supplier={}  #2)
products_under_10_inv={}  # 3)


for product_row in range(2,product_list.max_row+1):
    supplier_name=product_list.cell(product_row,4).value  #this will give us a supplier name for each row
    inventory=product_list.cell(product_row,2).value    #this will give us the inventory available in each row
    price = product_list.cell(product_row, 3).value     #this will give us the prices of each rowe 
    product_num = product_list.cell(product_row, 1).value  # this will give us the product number of each row
    inventor_price=product_list.cell(product_row,5)    #this will give us the access to the new column

    # calculation of number of products per supplier
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name]+=1      #inside brackets is key and rhs is value ie no of products per suppliers
    else:
        product_per_supplier[supplier_name]=1     #  print("addding new supplier to the dictionary") #just to visualise the addition of supplier to the dictionary

    
    # calculation of total value of invemtory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic for calculation of produtcs having inventory less than 10
    if inventory<10:
        products_under_10_inv[product_row]=int(inventory)
    
    # 5) total inventory price
    inventor_price.value= inventory * price
    

# print(product_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
print(inventor_price)
inv_file.save("updated.xlsx")  #saving file to reflect changes 


